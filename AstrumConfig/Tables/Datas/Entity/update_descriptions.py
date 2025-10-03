import openpyxl

# 更新 EntityBase.xlsx 的描述
wb = openpyxl.load_workbook('#EntityBase.xlsx')
ws = wb.active

# 设置中文描述行
ws['A4'] = '##desc'
ws['B4'] = '实体ID'
ws['C4'] = '模型ID'
ws['D4'] = '静止动作'
ws['E4'] = '走路动作'
ws['F4'] = '跑步动作'
ws['G4'] = '跳跃动作'
ws['H4'] = '出生动作'
ws['I4'] = '死亡动作'

wb.save('#EntityBase.xlsx')
print('#EntityBase.xlsx 描述已更新为中文')

# 更新 Action.xlsx 的描述
wb2 = openpyxl.load_workbook('#Action.xlsx')
ws2 = wb2.active

# 设置中文描述行
ws2['A4'] = '##desc'
ws2['B4'] = '动作ID'
ws2['C4'] = '动作名称'
ws2['D4'] = '动作类型'
ws2['E4'] = '持续时间'

wb2.save('#Action.xlsx')
print('#Action.xlsx 描述已更新为中文')

# 更新 EntityModel.xlsx 的描述
wb3 = openpyxl.load_workbook('#EntityModel.xlsx')
ws3 = wb3.active

# 设置中文描述行
ws3['A4'] = '##desc'
ws3['B4'] = '模型ID'
ws3['C4'] = '模型名称'
ws3['D4'] = '模型路径'
ws3['E4'] = '模型类型'

wb3.save('#EntityModel.xlsx')
print('#EntityModel.xlsx 描述已更新为中文')




